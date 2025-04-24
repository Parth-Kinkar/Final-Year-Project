from rest_framework import generics, permissions, status, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.shortcuts import get_object_or_404
from .models import CustomUser, Project, Student, Department
from .serializers import UserSerializer, ProjectSerializer, StudentSerializer, DepartmentSerializer
from django_filters.rest_framework import DjangoFilterBackend
import random
import string
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.http import HttpResponse
from .excel_generator import generate_excel_template  # Assuming you have this function in a separate file
import openpyxl
from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.parsers import MultiPartParser

def download_excel_template(request):
    return generate_excel_template()

# User Detail View (Updated to include profile_photo and full name)
class UserDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        student_profile = getattr(user, "student", None)  # Check if user is a student

        return Response({
            "id": user.id, 
            "username": user.username,
            "full_name": student_profile.full_name if student_profile else user.get_full_name(),
            "email": user.email,
            "user_type": user.user_type,
            "profile_photo": request.build_absolute_uri(user.profile_photo.url) if user.profile_photo else None,
            "contact_number": student_profile.contact_number if student_profile else None,
            "skills": student_profile.skills if student_profile else None,
            "interested_technologies": student_profile.interested_technologies if student_profile else None,
            "projects_created": [
                {"id": project.id, "title": project.title, "description": project.description}
                for project in Project.objects.filter(creators=user)
            ] if student_profile else [],
            "github": student_profile.github if student_profile else None,  # Include GitHub link
            "linkedin": student_profile.linkedin if student_profile else None,  # Include LinkedIn link
            "portfolio": student_profile.portfolio if student_profile else None,  # Include Portfolio link
            "resume": request.build_absolute_uri(student_profile.resume.url) if student_profile and student_profile.resume else None,  # Include Resume
        })

# Register View (Unchanged)
class RegisterView(generics.CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.AllowAny]

# Login View (Unchanged)
class LoginView(generics.GenericAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        try:
            user = CustomUser.objects.get(username=username)
        except CustomUser.DoesNotExist:
            return Response({'error': 'Invalid credentials'}, status=400)

        if user.check_password(password):
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user_type': user.user_type,
                "user": {  # ✅ Ensure `user` object is correctly formatted
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "user_type": user.user_type,
                }
            })
        return Response({'error': 'Invalid credentials'}, status=400)

# Student List View (Updated to use StudentSerializer)
class StudentListView(generics.ListAPIView):
    queryset = Student.objects.select_related('user', 'department').all()  # Optimize queries
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]

# Department List View (New, if you need to fetch department options in frontend)
class DepartmentListView(generics.ListAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [permissions.AllowAny]

class DepartmentChoicesView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        # Fetch DEPARTMENT_CHOICES dynamically from the model
        choices = [{'value': value, 'label': label} for value, label in Department.DEPARTMENT_CHOICES]
        return Response(choices)
    
class DepartmentCreateView(generics.CreateAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]  # Ensure only logged-in users can add
    
class DepartmentDeleteView(generics.DestroyAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]

# Project Views (Unchanged)
class ProjectListView(generics.ListAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]

class ProjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_context(self):
        return {"request": self.request}  # Pass request to serializer

class ProjectCreateView(APIView):
    permission_classes = [IsAuthenticated]  

    def post(self, request, *args, **kwargs):
        data = request.data
        serializer = ProjectSerializer(data=data)

        if serializer.is_valid():
            project = serializer.save()

            # Convert creator IDs to User objects
            creators_data = data.get('creators', [])
            students = CustomUser.objects.filter(id__in=creators_data, user_type='student')
            project.creators.set(students)  
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Teacher List View (Unchanged)
class TeacherListView(generics.ListAPIView):
    queryset = CustomUser.objects.filter(user_type='teacher')
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

# Fetch Students by Year and Department
class FilteredStudentListView(generics.ListAPIView):
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Student.objects.select_related("user", "department").all()
        year = self.request.query_params.get("year")
        department = self.request.query_params.get("department")

        if year:
            queryset = queryset.filter(year=year)
        if department:
            queryset = queryset.filter(department__name=department)

        return queryset

#Bookmarked projects view:
class BookmarkProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, project_id):
        student = get_object_or_404(Student, user=request.user)
        project = get_object_or_404(Project, id=project_id)

        if project in student.bookmarked_projects.all():
            student.bookmarked_projects.remove(project)  # Remove if already bookmarked
            message = "Removed from bookmarks."
        else:
            student.bookmarked_projects.add(project)  # Add if not bookmarked
            message = "Added to bookmarks."

        student.save()
        return Response({"message": message, "bookmarked_projects": list(student.bookmarked_projects.values_list('id', flat=True))})

# Edit profile View
class EditUserProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        user = request.user  # Get the logged-in user
        data = request.data  # Data from the frontend
        
        try:
            # Update fields in the CustomUser model
            user.email = data.get("email", user.email)
            user.save()

            # Update fields in the Student model (if user is a student)
            student_profile = getattr(user, "student", None)  # Get the related Student profile
            if student_profile:
                student_profile.full_name = data.get("full_name", student_profile.full_name)
                student_profile.skills = data.get("skills", student_profile.skills)
                student_profile.interested_technologies = data.get(
                    "interested_technologies", student_profile.interested_technologies
                )
                student_profile.contact_number = data.get("contact_number", student_profile.contact_number)
                student_profile.resume = data.get("resume", student_profile.resume)
                student_profile.github = data.get("github", student_profile.github)
                student_profile.linkedin = data.get("linkedin", student_profile.linkedin)
                student_profile.portfolio = data.get("portfolio", student_profile.portfolio)

                student_profile.save()

            return Response({"message": "Profile updated successfully!"}, status=status.HTTP_200_OK)
        except Exception as e:
            # Log the error for debugging
            print(f"Error updating profile: {str(e)}")
            return Response({"error": "Failed to update profile."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class SearchView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = request.query_params.get('q', '').strip()

        # Search matching profiles
        matching_profiles = Student.objects.filter(
        full_name__icontains=query
        ).values('user_id', 'full_name', 'user__username', 'user__profile_photo')

        # Search matching projects
        matching_projects = Project.objects.filter(
            title__icontains=query
        ).values('id', 'title', 'description', 'creators__username')

        return Response({
            'profiles': list(matching_profiles),
            'projects': list(matching_projects),
        })

class StudentSearchView(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    
    # Enable filtering by department and year
    filterset_fields = ['department__id', 'year']
    
    # Enable searching by name, roll number, and email
    search_fields = ['full_name', 'roll_number', 'user__email']

class StudentUpdateView(generics.UpdateAPIView):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user_id = self.kwargs.get("user_id")  # Fetching user ID from URL
        return Student.objects.get(user__id=user_id)

class StudentDeleteView(generics.DestroyAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        user_id = self.kwargs.get("user_id")  # Fetch user ID from URL
        return Student.objects.get(user__id=user_id)

class StudentCreateView(generics.CreateAPIView):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        full_name = request.data.get("full_name")
        email = request.data.get("email")
        roll_number = request.data.get("roll_number")
        year = request.data.get("year")
        department_id = request.data.get("department_id")

        if not full_name or not email or not roll_number or not department_id:
            return Response({"error": "Missing required fields."}, status=status.HTTP_400_BAD_REQUEST)

        username = full_name.lower().replace(" ", ".") + str(random.randint(100, 999))
        default_password = "Student@123"

        user = CustomUser.objects.create_user(
            username=username,
            password=default_password,
            email=email,
            user_type="student"
        )

        department = Department.objects.get(id=department_id)
        Student.objects.create(user=user, roll_number=roll_number, department=department, year=year)

        # ✅ Since `create()` returns the final response, it will NOT be overridden by serializer logic
        return Response({
            "message": "Student created successfully!",
            "username": username,
            "password": default_password
        }, status=status.HTTP_201_CREATED)

DEFAULT_PASSWORD = "Student@123"
# class UploadExcelView(APIView):
#     def post(self, request):
#         file = request.FILES.get("excel_file")
#         department_id = request.data.get("department_id")
#         year = request.data.get("year")

#         if not file or not department_id or not year:
#             return Response({"error": "File, Year, or Department missing"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             wb = openpyxl.load_workbook(file)
#             ws = wb.active
#             students = []
#             department = Department.objects.get(id=department_id)

#             for row in ws.iter_rows(min_row=2, values_only=True):  # Skip headers
#                 full_name, email, roll_number = row

#                 if not full_name or not email or not roll_number:
#                     continue  # Skip invalid rows

#                 # ✅ Generate username dynamically
#                 username = full_name.lower().replace(" ", ".") + str(random.randint(100, 999))

#                 # ✅ Create user and student entry
#                 user = CustomUser.objects.create_user(
#                     username=username,
#                     password=DEFAULT_PASSWORD,
#                     email=email,
#                     user_type="student"
#                 )

#                 Student.objects.create(user=user, roll_number=roll_number, department=department, year=year)

#                 students.append({
#                     "full_name": full_name,
#                     "email": email,
#                     "roll_number": roll_number,
#                     "username": username,
#                     "password": DEFAULT_PASSWORD
#                 })

#             return Response({"students": students}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response({"error": f"Invalid file format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
class UploadExcelView(APIView):
    def post(self, request):
        file = request.FILES.get("excel_file")
        department_id = request.data.get("department_id")
        year = request.data.get("year")

        if not file or not department_id or not year:
            return Response({"error": "File, Year, or Department missing"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            students = []

            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip headers
                full_name, email, roll_number = row

                if not full_name or not email or not roll_number:
                    continue  # Skip invalid rows

                # ✅ Generate username dynamically but DON'T insert yet
                username = full_name.lower().replace(" ", ".") + str(random.randint(100, 999))

                students.append({
                    "full_name": full_name,
                    "email": email,
                    "roll_number": roll_number,
                    "username": username,
                    "password": "Student@123"  # Common password
                })

            return Response({"students": students}, status=status.HTTP_200_OK)  # ✅ Just return preview data

        except Exception as e:
            return Response({"error": f"Invalid file format: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


class ConfirmUploadView(APIView):
    def post(self, request):
        students_data = request.data.get("students")

        if not students_data:
            return Response({"error": "No student data provided"}, status=status.HTTP_400_BAD_REQUEST)

        created_students = []
        for student in students_data:
            full_name = student.get("full_name")
            email = student.get("email")
            roll_number = student.get("roll_number")
            username = student.get("username")
            department_id = request.data.get("department_id")
            year = request.data.get("year")

            if not full_name or not email or not roll_number or not department_id or not year:
                continue  # Skip invalid entries
            if CustomUser.objects.filter(email=email).exists():
                return Response({"error": f"Email {email} is already in use"}, status=status.HTTP_400_BAD_REQUEST)

            department = Department.objects.get(id=department_id)
            user = CustomUser.objects.create_user(
                username=username,
                password="Student@123",  # Default password
                email=email,
                user_type="student"
            )

            Student.objects.create(user=user, roll_number=roll_number, department=department, year=year)
            created_students.append(student)

        return Response({"message": "Students added successfully!", "students": created_students}, status=status.HTTP_201_CREATED)